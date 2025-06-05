from flask import Flask, render_template, request, redirect, url_for, jsonify, session
import openpyxl
import os
import json
import requests
from datetime import datetime

app = Flask(__name__)
app.secret_key = "your_secret_key"  # Needed for session

EXCEL_FILE_PATH = "users.xlsx"

# Create Excel file if it doesn't exist
if not os.path.exists(EXCEL_FILE_PATH):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "users"
    ws1.append(["username", "password"])
    ws2 = wb.create_sheet(title="messages")
    ws2.append(["username", "sender", "message", "timestamp"])
    wb.save(EXCEL_FILE_PATH)

def load_workbook():
    return openpyxl.load_workbook(EXCEL_FILE_PATH)

def log_message(username, sender, message):
    wb = load_workbook()
    if "messages" not in wb.sheetnames:
        ws = wb.create_sheet("messages")
        ws.append(["username", "sender", "message", "timestamp"])
    ws = wb["messages"]
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([username, sender, message, timestamp])
    wb.save(EXCEL_FILE_PATH)

@app.route("/")
def home():
    return render_template("home.html")

@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        wb = load_workbook()
        ws = wb["users"]

        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[0] == username:
                return "Username already taken."

        ws.append([username, password])
        wb.save(EXCEL_FILE_PATH)
        return redirect(url_for("login"))
    return render_template("signup.html")

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        wb = load_workbook()
        ws = wb["users"]

        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            if row[0] == username and row[1] == password:
                session["username"] = username
                return redirect(url_for("index"))

        return "Invalid username or password!"
    return render_template("login.html")

@app.route("/index")
def index():
    if "username" not in session:
        return redirect(url_for("login"))
    return render_template("index.html")

@app.route("/chat", methods=["POST"])
def chat():
    if "username" not in session:
        return jsonify({"reply": "You are not logged in."})

    username = session["username"]
    user_message = request.json["message"]
    log_message(username, "user", user_message)

    payload = {
        "model": "llama3",
        "messages": [
            {
                "role": "system",
                "content": (
                    "You're Tomo, a cheerful, emotionally supportive AI friend. "
                    "Talk casually like a real friend, using warmth, comfort, and empathy. "
                    "Sprinkle in emojis and encouragement, avoid robotic language."
                )
            },
            {
                "role": "user",
                "content": user_message
            }
        ],
        "stream": True
    }

    response = requests.post(
        "http://localhost:11434/api/chat",
        headers={"Content-Type": "application/json"},
        data=json.dumps(payload),
        stream=True
    )

    reply = ""
    for line in response.iter_lines():
        if line:
            data = json.loads(line.decode("utf-8"))
            if "message" in data and "content" in data["message"]:
                reply += data["message"]["content"]

    log_message(username, "Tomo", reply)
    return jsonify({"reply": reply})

@app.route("/logout")
def logout():
    session.pop("username", None)
    return redirect(url_for("login"))

if __name__ == "__main__":
    app.run(debug=True)
